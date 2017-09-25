import pandas as pd
import numpy as np
import datetime as dt
import xlsxwriter
import csv
from openpyxl import load_workbook
import os
import sys

# Date Variables
def ds_start():
    # List all spreadsheets that aren't the main DS workbook
    file = [each for each in os.listdir(cwd) if each.endswith('-2017.xlsx') and 'Daily Sales' not in each][0]
    if len(files) != 1:
        error_file('More than one file found')
        sys.exit()
    else:
        create_sales_df(filename)

def create_sales_df(file):
    # Import Sales
    sales_date = filename
    sales_xl = pd.ExcelFile(sales_date + '.xlsx')
    sales_df = sales_xl.parse('Sheet1')
    sales_df = sales_df[['GL Account #', 'Site', 'End Date', '(Item) Name', 'Amount']]

    # Remove Amounts that equal zero
    sales_df['Amount'] = sales_df['Amount'].fillna(0)
    sales_df = sales_df[sales_df['Amount'].isin([0])==False]
    
    # Import GL Codes
    gl_xl = pd.ExcelFile('dailysales.xlsx')
    gl_df_base = gl_xl.parse('Base', index_col='(Item) Name')
    gl_df_base = gl_df_base.loc[~gl_df_base.index.duplicated(keep='first')]
    gl_df_site = gl_xl.parse('Site', index_col='Site')

    # Match Description to GL Codes (VLOOKUP)
    sales_df['GL Account #'] = sales_df['(Item) Name'].map(gl_df_base['GL Account #'].astype(str))
    sales_df['PWC'] = sales_df['(Item) Name'].map(gl_df_base['PWC'].astype(str))
    sales_df['Site_ID'] = sales_df['Site'].map(gl_df_site['Site_ID'].astype(str))
    
    # Items that aren't GL coded - add to dailysales
    sales_df_miss = sales_df[sales_df['GL Account #'].isnull()][['(Item) Name', 'GL Account #']]
    sales_df_miss = sales_df_miss.set_index('(Item) Name').fillna('-')
    if not sales_df_miss.empty:
        print('Site missing GL Codes')
        gl_df_base = gl_df_base.append(sales_df_miss).reset_index()
        gl_df_base = gl_df_base.loc[~gl_df_base.index.duplicated(keep='first')]
        
        book = load_workbook('dailysales.xlsx')
        writer = pd.ExcelWriter('dailysales.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        gl_df_base.to_excel(writer, 'Base', header=True, index=False)
        writer.save()
    
    # Sites that aren't GL coded - add to dailysales
    sales_df_miss = sales_df[sales_df['Site_ID'].isnull()][['Site', 'Site_ID']]
    sales_df_miss = sales_df_miss.set_index('Site').fillna('-')
    if not sales_df_miss.empty:
        gl_df_site = gl_df_site.append(sales_df_miss).reset_index()
        
        book = load_workbook('dailysales.xlsx')
        writer = pd.ExcelWriter('dailysales.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        gl_df_base.to_excel(writer, 'Site', header=True, index=False)
        writer.save()

    # Some formatting
    sales_df['End Date'] = sales_df['End Date'].dt.strftime('%-m/%-d/%y')
    dates = sales_df['End Date'].unique().tolist()
    site_name = sales_df.Site.unique().tolist()[0]
    site_id = sales_df.Site_ID.unique().tolist()[0]

    # Make sure Amount col reconciles
    for date in dates:
        sales_df_date = sales_df[sales_df['End Date']==date]
        
        # Enter new data for 2164 account
        cols = ['GL Account #', 'Site', 'End Date', '(Item) Name', 'Amount', 'Site_ID']
        sales_df_date_2164 = sales_df_date[sales_df_date['GL Account #']=='2164']
        if not sales_df_date_2164.empty and round(sales_df_date_2164['Amount'].sum(),2) != 0:
            new_2164 = round(sales_df_date_2164['Amount'].sum(),2)
            new_2165 = round(sales_df_date_2164['Amount'].sum(),2) * -1
            
            df_2164 = pd.DataFrame([['2164', site_name, date, 'GSR B1G1 Sold', new_2164, site_id]],columns=cols)
            df_2165 = pd.DataFrame([['2165', site_name, date, 'GSR B1G1 Sold Dscnt', new_2165, site_id]],columns=cols)
            df_entries = df_2164.append(df_2165, ignore_index=True)
            sales_df = sales_df.append(df_entries, ignore_index=True)
        
        # PWC
        sales_df_date_PWC = sales_df_date[sales_df_date['PWC']=='x']
        if not sales_df_date_PWC.empty and round(sales_df_date_PWC['Amount'].sum(),2) != 0:
            new_1201 = round(sales_df_date_PWC['Amount'].sum() * 0.2,2)
            new_2163 = round(sales_df_date_PWC['Amount'].sum() * 0.8,2)
            new_2162 = (new_1201 + new_2163) * -1

            df_1201 = pd.DataFrame([['1201', site_name, date, 'Accounts Receivable', new_1201, site_id]],columns=cols)
            df_2163 = pd.DataFrame([['2163', site_name, date, 'GSR PC - Discount', new_2163, site_id]],columns=cols)
            df_2162 = pd.DataFrame([['2162', site_name, date, 'GSR PC - Dscnt', new_2162, site_id]],columns=cols)
            df_entries = df_1201.append(df_2163, ignore_index=True)
            df_entries = df_entries.append(df_2162, ignore_index=True)
            
            sales_df = sales_df.append(df_entries, ignore_index=True)
        
        # Does this day reconcile?
        sales_df_rec = round(sales_df_date['Amount'].sum(),2)
        if sales_df_rec != 0:
            error_file(date + ' - off by $' + str(sales_df_rec))
    
    # Create GL Upload Columns
    sales_df['RECORD'] = 'GLT'
    sales_df['ACCOUNT'] = sales_df['Site_ID'] + '-' + sales_df['GL Account #'] + '.000'
    sales_df['ACCNTG DATE'] = sales_df['End Date']
    sales_df['JOURNAL'] = 10
    sales_df['REF 1'] = ''
    sales_df['REF 2'] = ''
    sales_df['DESCRIPTION'] = sales_df['(Item) Name'].str[0:30]
    sales_df['DEBIT'] = np.where(sales_df['Amount'] < 0, sales_df['Amount'] * -1,0)
    sales_df['CREDIT'] = np.where(sales_df['Amount'] > 0, sales_df['Amount'],0)
    sales_df['ACCRUAL OR CASH'] = 1
    
    # Finish Magic
    update_workbook(now, sales_df)

def update_workbook(now, sales_df):
    # Date Setup
    month_names = {'1': 'January', '2': 'February', '3': 'March',
            '4': 'April', '5': 'May', '6': 'June',
            '7': 'July', '8': 'August', '9': 'September',
            '10': 'October', '11': 'November', '12': 'December'}
    
    # Index Site Name
    sites = sales_df.Site.unique().tolist()
    if len(sites) == 1:
        site_name = sites[0]
    else:
        error_file('Number of Sites Error: ' + str(sites))

    # Account Summary
    sales_df_sum = sales_df[sales_df['GL Account #'] != '1099'].groupby('ACCOUNT')['Amount'].sum()
    sales_df_1099 = sales_df[sales_df['GL Account #'] == '1099'].groupby('(Item) Name')['Amount'].sum()
    sales_df_sum = sales_df_sum.append(sales_df_1099)

    # Slice GL Upload Format
    upload_format = ['RECORD', 'ACCOUNT', 'ACCNTG DATE', 'JOURNAL', 'REF 1', 'REF 2', 'DESCRIPTION', 'DEBIT', 'CREDIT', 'ACCRUAL OR CASH']
    sales_df = sales_df[upload_format]

    # DS Filename/Path
    ds_filename = 'Daily Sales Entry - ' + site_name + ' - ' + now['month-name'] + ' ' + now['year'] + '.xlsx'
    ds_filepath = cwd + '/' + ds_filename
    
    # Create DS File if it doesn't exist
    if os.path.exists(ds_filepath) == False:
        workbook = xlsxwriter.Workbook(ds_filename)
        worksheet = workbook.add_worksheet('1')
        workbook.close()
    
    # Save each Date as a Sheet
    dates = sales_df['ACCNTG DATE'].unique().tolist()
    for date in dates:
        print(date)
        day = date.split('/')[1]
        # Copy old Workbook
        book = load_workbook(ds_filename)
        writer = pd.ExcelWriter(ds_filename, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        # Save Workbook with new Data
        sales_df_sheet = sales_df[sales_df['ACCNTG DATE']==date]
        sales_df_sheet.to_excel(writer, day, header=True, index=False)
        writer.save()

    # Create Summary Tab
    book = load_workbook(ds_filename)
    writer = pd.ExcelWriter(ds_filename, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    sales_df_sum.reset_index().to_excel(writer, 'Summary', header=True, index=False)
    writer.save()
    
    create_upload_file(ds_filename)

def create_upload_file(filename):
    # Load workbook
    xls = pd.ExcelFile(filename)
    # Create empty DF
    upload_df = pd.DataFrame()
    # Fill DF
    for ws in xls.sheet_names:
        if ws != 'Summary':
            upload_df = upload_df.append(xls.parse(ws), ignore_index=True)
    # Save DF as CSV
    upload_df.to_csv('{}.csv'.format(filename), header=True, index=False)


def error_file(error):
    error = str(error).replace('/','-')
    try:
        file = open(error + '.txt', 'w')
        file.write(error)
        file.close()
    except:
        print('Could not create error file: ' + error)

# Main
if __name__ == '__main__':
    cwd = os.getcwd()
    ds_start()
